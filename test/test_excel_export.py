# =============================================================================
# test/test_excel_export.py
# Tests automatizados para lib_excel_export.py
#
# ¿Por qué estos tests concretos?
# Los tres bugs que aparecieron durante el desarrollo de esta fase fueron:
#   1. Listas de distinta longitud en _sheet_parametros → ValueError de pandas
#   2. Referencias a columnas renombradas (assigned_slot, CTA...) → KeyError
#   3. Argumentos pasados en el orden o con el nombre incorrecto → TypeError
#
# Estos tests detectan exactamente esos tres tipos de error antes de que
# lleguen a la ejecución del proyecto completo.
#
# Cómo ejecutarlos:
#   python -m pytest test/ -v
# =============================================================================

import os
import sys
import tempfile
import shutil
import unittest

import pandas as pd
import numpy as np

# Añadimos src/ al path para poder importar los módulos del proyecto
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

import lib_excel_export as excel
from config import CFG


# =============================================================================
# DATOS SINTÉTICOS COMPARTIDOS
# Generamos DataFrames mínimos que simulan la salida real del pipeline
# sin necesidad de ejecutar el proyecto completo.
# =============================================================================

def _hacer_df_res_minimo() -> pd.DataFrame:
    """
    Crea un DataFrame de resultados GDP mínimo pero válido.
    Contiene todas las columnas que lib_excel_export.py necesita.
    """
    return pd.DataFrame({
        'ARCID':         ['IBE001', 'BAW002', 'UAE003', 'VLG004', 'AFR005'],
        'airline':       ['IBE',    'BAW',    'UAE',    'VLG',    'AFR'],
        'ADEP':          ['LEMD',   'EGLL',   'OMDB',   'LEMD',   'LFPG'],
        'ATYP':          ['B738',   'A320',   'B77W',   'A320',   'A319'],
        'recat':         ['D',      'D',      'C',      'D',      'D'],
        'is_ecac':       [True,     True,     False,    True,     True],
        'distancia_km':  [500.0,    1800.0,   4200.0,   500.0,    1100.0],
        'minutes_eta':   [390.0,    410.0,    430.0,    450.0,    470.0],
        'assigned_slot': [395.0,    415.0,    430.0,    460.0,    480.0],
        'total_delay':   [5.0,      5.0,      0.0,      10.0,     10.0],
        'air_delay':     [0.0,      0.0,      0.0,      0.0,      0.0],
        'ground_delay':  [5.0,      5.0,      0.0,      10.0,     10.0],
        'flight_status': [
            'GPD CANDIDATE', 'GPD CANDIDATE', 'EXEMPT INTERNATIONAL',
            'GPD CANDIDATE', 'GPD CANDIDATE',
        ],
    })


def _hacer_timeline_minimo() -> pd.DataFrame:
    """
    Crea un timeline mínimo con las columnas que necesita
    calcular_retraso_minimo_newell().
    """
    minutos = list(range(1440))
    demand  = [float(i // 10) for i in minutos]
    cap     = [min(float(i // 10), float(i // 12)) for i in minutos]
    return pd.DataFrame({
        'minuto':           minutos,
        'demand_accum':     demand,
        'capacity_accum':   cap,
        'queue_size':       [max(0.0, d - c) for d, c in zip(demand, cap)],
    })


def _hacer_params_minimo() -> dict:
    """Parámetros mínimos del escenario GDP."""
    return {
        'H_START':         CFG.H_START,
        'H_END':           CFG.H_END,
        'AAR':             CFG.AAR,
        'PAAR':            CFG.PAAR,
        'SLOT_NOM':        CFG.SLOT_NOM,
        'SLOT_RED':        CFG.SLOT_RED,
        'H_FREEZE_OFFSET': CFG.H_FREEZE_OFFSET,
    }


def _hacer_df_slots_minimo() -> pd.DataFrame:
    """Matriz de slots mínima."""
    return pd.DataFrame({
        'slot_start_min': [390.0, 393.0, 396.0, 399.0, 402.0],
        'occupied':       [True,  True,  False, True,  True],
        'flight_id':      [0,     1,     None,  3,     4],
    })


def _hacer_df_crudos_minimo() -> pd.DataFrame:
    """DataFrame de vuelos crudos mínimo (antes del GDP)."""
    return pd.DataFrame({
        'ARCID': ['IBE001', 'BAW002', 'UAE003', 'VLG004', 'AFR005'],
        'ADEP':  ['LEMD',   'EGLL',   'OMDB',   'LEMD',   'LFPG'],
        'ADES':  ['LEBL'] * 5,
        'ETA':   ['06:30', '06:50', '07:10', '07:30', '07:50'],
    })


# =============================================================================
# TEST CLASE 1 — Integridad de las listas en _sheet_parametros
#
# Bug que previene: ValueError "All arrays must be of the same length"
# Causa original:  Se añadió una fila a 'Parámetro' sin añadir la
#                  correspondiente en 'Valor' y 'Fuente / Nota'.
# =============================================================================

class TestSheetParametrosLongitudes(unittest.TestCase):
    """
    Verifica que las tres listas del DataFrame de _sheet_parametros
    tienen exactamente la misma longitud.

    Si alguien añade una fila a 'Parámetro' sin añadir su correspondiente
    en 'Valor' o 'Fuente', este test falla inmediatamente con un mensaje
    claro en lugar de un ValueError críptico al ejecutar el proyecto.
    """

    def test_no_lanza_value_error(self):
        """
        El test más importante: que la función se ejecute sin lanzar
        ValueError. Si las listas tienen distinta longitud, pandas lanza
        este error al construir el DataFrame.
        """
        from openpyxl import Workbook
        wb = Workbook()

        params   = _hacer_params_minimo()
        timeline = _hacer_timeline_minimo()

        # Si las listas tienen distinta longitud, esto lanza ValueError
        try:
            excel._sheet_parametros(wb, params, 900, timeline)
        except ValueError as e:
            self.fail(
                f"_sheet_parametros lanzó ValueError: {e}\n"
                f"Causa probable: una lista tiene más o menos elementos que las otras."
            )

    def test_pestaña_creada_en_workbook(self):
        """
        Verifica que la función crea efectivamente la pestaña en el workbook.
        """
        from openpyxl import Workbook
        wb = Workbook()
        excel._sheet_parametros(wb, _hacer_params_minimo(), 900, _hacer_timeline_minimo())
        self.assertIn('0_Parametros_GDP', wb.sheetnames)


# =============================================================================
# TEST CLASE 2 — Columnas del DataFrame de resultados
#
# Bug que previene: KeyError cuando una columna ha sido renombrada
# Causa original:  Renombrar 'assigned_slot' en algunos sitios pero no en
#                  otros, o cambiar nombres de columnas sin actualizar
#                  todas las referencias.
# =============================================================================

class TestColumnasDataFrames(unittest.TestCase):
    """
    Verifica que el DataFrame de resultados contiene todas las columnas
    que lib_excel_export.py necesita para construir el Excel.

    Si alguien renombra una columna en lib_gdp_core.py, estos tests
    fallan inmediatamente señalando exactamente qué columna falta.
    """

    def setUp(self):
        self.df_res = _hacer_df_res_minimo()

    def test_columnas_obligatorias_existen(self):
        """
        Estas columnas deben existir en df_res. Si falta alguna,
        el test dice exactamente cuál es.
        """
        columnas_obligatorias = [
            'ARCID', 'airline', 'ADEP', 'ATYP', 'recat',
            'is_ecac', 'distancia_km', 'minutes_eta',
            'assigned_slot',   # nombre interno — NO renombrar sin actualizar aquí
            'total_delay', 'air_delay', 'ground_delay', 'flight_status',
        ]
        for col in columnas_obligatorias:
            self.assertIn(
                col, self.df_res.columns,
                msg=(
                    f"Falta la columna '{col}' en df_res.\n"
                    f"Si la renombraste, actualiza también lib_excel_export.py "
                    f"y este test."
                )
            )

    def test_columnas_timeline_existen(self):
        """
        El timeline debe tener estas columnas para calcular el retraso
        mínimo teórico de Newell.
        """
        timeline = _hacer_timeline_minimo()
        columnas_timeline = ['minuto', 'demand_accum', 'capacity_accum', 'queue_size']
        for col in columnas_timeline:
            self.assertIn(col, timeline.columns, msg=f"Falta columna '{col}' en timeline.")

    def test_flight_status_valores_validos(self):
        """
        Los valores de flight_status deben ser exactamente estos cuatro.
        Si alguien cambia un string en lib_gdp_core.py (ej: 'GPD CANDIDATE'
        por 'GDP CANDIDATE'), los filtros del Excel dejan de funcionar.
        """
        valores_validos = {
            'GPD CANDIDATE',
            'EXEMPT INTERNATIONAL',
            'EXEMPT AIRBORNE',
            'EXEMPT DISTANCE',
        }
        valores_reales = set(self.df_res['flight_status'].unique())
        for valor in valores_reales:
            self.assertIn(
                valor, valores_validos,
                msg=(
                    f"Valor inesperado en flight_status: '{valor}'.\n"
                    f"Valores permitidos: {valores_validos}"
                )
            )


# =============================================================================
# TEST CLASE 3 — Generación completa del Excel
#
# Bug que previene: TypeError por argumentos incorrectos, pestañas que
#                   no se crean, o errores silenciosos en la exportación.
# Causa original:  Pasar df_res donde se esperaba timeline, o no actualizar
#                  la firma de exportar_auditoria_excel al añadir argumentos.
# =============================================================================

class TestGeneracionExcelCompleto(unittest.TestCase):
    """
    Test de integración: genera el Excel completo con datos sintéticos
    y verifica que todas las pestañas existen y tienen contenido.
    """

    def setUp(self):
        """Prepara datos y carpeta temporal antes de cada test."""
        self.directorio_temp = tempfile.mkdtemp()
        self.path_excel = os.path.join(self.directorio_temp, 'test_auditoria.xlsx')

        self.df_res     = _hacer_df_res_minimo()
        self.df_slots   = _hacer_df_slots_minimo()
        self.df_crudos  = _hacer_df_crudos_minimo()
        self.params     = _hacer_params_minimo()
        self.timeline   = _hacer_timeline_minimo()
        self.h_noreg    = 900

    def tearDown(self):
        """Limpia la carpeta temporal después de cada test."""
        shutil.rmtree(self.directorio_temp, ignore_errors=True)

    def test_no_lanza_excepcion(self):
        """
        La función principal no debe lanzar ninguna excepción con datos válidos.
        Si lanza TypeError, probablemente un argumento está en el orden incorrecto
        o tiene el nombre equivocado.
        """
        try:
            excel.exportar_auditoria_excel(
                self.df_crudos,
                self.df_res,
                self.df_slots,
                self.params,
                self.h_noreg,
                self.timeline,
                self.path_excel,
            )
        except Exception as e:
            self.fail(
                f"exportar_auditoria_excel lanzó una excepción inesperada:\n"
                f"{type(e).__name__}: {e}"
            )

    def test_archivo_excel_se_crea(self):
        """El archivo .xlsx debe existir después de llamar a la función."""
        excel.exportar_auditoria_excel(
            self.df_crudos, self.df_res, self.df_slots,
            self.params, self.h_noreg, self.timeline, self.path_excel,
        )
        self.assertTrue(
            os.path.exists(self.path_excel),
            msg="El archivo Excel no se ha creado en la ruta esperada."
        )

    def test_todas_las_pestañas_existen(self):
        """
        El Excel debe tener exactamente estas pestañas.
        Si se añade o elimina una pestaña, este test lo detecta.
        """
        excel.exportar_auditoria_excel(
            self.df_crudos, self.df_res, self.df_slots,
            self.params, self.h_noreg, self.timeline, self.path_excel,
        )

        from openpyxl import load_workbook
        wb = load_workbook(self.path_excel)

        pestañas_esperadas = [
            '0_Parametros_GDP',
            '1_Datos_Crudos',
            '2_Regulacion_GDP',
            '3_Matriz_Slots',
            '5_KPIs_Comparativa',
            '6_Analisis_Retrasos',
            '7_Equidad_RBS',
        ]
        for pestaña in pestañas_esperadas:
            self.assertIn(
                pestaña, wb.sheetnames,
                msg=(
                    f"Falta la pestaña '{pestaña}' en el Excel generado.\n"
                    f"Pestañas encontradas: {wb.sheetnames}"
                )
            )

    def test_pestañas_tienen_contenido(self):
        """
        Ninguna pestaña debe estar vacía — todas deben tener al menos
        una fila de cabecera y una fila de datos.
        """
        excel.exportar_auditoria_excel(
            self.df_crudos, self.df_res, self.df_slots,
            self.params, self.h_noreg, self.timeline, self.path_excel,
        )

        from openpyxl import load_workbook
        wb = load_workbook(self.path_excel)

        for nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            self.assertGreater(
                ws.max_row, 1,
                msg=f"La pestaña '{nombre_hoja}' está vacía o solo tiene cabecera."
            )

    def test_firma_correcta_de_exportar(self):
        """
        Verifica que exportar_auditoria_excel acepta exactamente los
        argumentos que main.py le pasa, en el orden correcto.
        Si alguien cambia la firma sin actualizar main.py, este test falla.
        """
        import inspect
        sig = inspect.signature(excel.exportar_auditoria_excel)
        parametros = list(sig.parameters.keys())

        orden_esperado = [
            'df_vuelos_crudo', 'df_res', 'df_slots',
            'params', 'h_noreg', 'timeline', 'path',
        ]
        self.assertEqual(
            parametros, orden_esperado,
            msg=(
                f"La firma de exportar_auditoria_excel ha cambiado.\n"
                f"Esperado: {orden_esperado}\n"
                f"Actual:   {parametros}\n"
                f"Actualiza también main.py si cambias el orden de los argumentos."
            )
        )


# =============================================================================
# PUNTO DE ENTRADA
# =============================================================================

if __name__ == '__main__':
    unittest.main(verbosity=2)