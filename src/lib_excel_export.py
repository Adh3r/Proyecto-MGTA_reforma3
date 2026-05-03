# =============================================================================
# src/lib_excel_export.py
# FASE 3: Construcción y formato del Excel maestro de auditoría (Versión Reporte Final)
# =============================================================================

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Importes de configuración (asegúrate de que coincidan con tus rutas)
try:
    from config import COST_AIR_MIN, COST_GND_MIN, FS_CANDIDATE
except ImportError:
    pass # Fallback por si ejecutas de forma aislada

# Colores Corporativos para el Reporte
COLOR_CABECERA_AZUL = "FF1F4E79"  # Azul Oscuro (Profesional)
COLOR_CABECERA_GHP  = "FF375623"  # Verde Oscuro (Para escenarios GHP/Optimizados)
COLOR_CABECERA_INT  = "FF7030A0"  # Púrpura (Para Intermodalidad)
COLOR_SECCION       = "FF2E75B6"
COLOR_FILA_ALTERNA  = "FFD9E1F2"
COLOR_FILA_BLANCA   = "FFFFFFFF"
FUENTE_BASE         = "Arial"

# =============================================================================
# UTILIDADES DE ESTILO
# =============================================================================

def _borde_fino():
    lado = Side(style="thin", color="FFD0D0D0")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _estilo_cabecera(color_fondo=COLOR_CABECERA_AZUL):
    return {
        'font':      Font(name=FUENTE_BASE, size=11, bold=True, color="FFFFFFFF"),
        'fill':      PatternFill("solid", fgColor=color_fondo),
        'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
        'border':    _borde_fino(),
    }

def _estilo_dato(fila_par=False, negrita=False):
    color = COLOR_FILA_ALTERNA if fila_par else COLOR_FILA_BLANCA
    return {
        'font':      Font(name=FUENTE_BASE, size=10, bold=negrita),
        'fill':      PatternFill("solid", fgColor=color),
        'alignment': Alignment(horizontal="left", vertical="center", wrap_text=False),
        'border':    _borde_fino(),
    }

def _autoajustar_columnas(ws, ancho_minimo=12, ancho_maximo=50):
    for col_cells in ws.columns:
        max_len = 0
        col_letra = get_column_letter(col_cells[0].column)
        for celda in col_cells:
            if celda.value is not None:
                max_len = max(max_len, len(str(celda.value)))
        ws.column_dimensions[col_letra].width = min(ancho_maximo, max(ancho_minimo, max_len * 1.1 + 2))

# =============================================================================
# MOTOR DE ESCRITURA CON AUTO-FORMATEO DE KPIs
# =============================================================================

def _escribir_dataframe_con_formato(ws, df, color_header=COLOR_CABECERA_AZUL):
    """Escribe un DF aplicando estilos y auto-detectando formatos de número según la columna."""
    # Cabeceras
    columnas = list(df.columns)
    for col_idx, nombre_col in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col_idx, value=nombre_col)
        estilos = _estilo_cabecera(color_header)
        celda.font = estilos['font']; celda.fill = estilos['fill']
        celda.alignment = estilos['alignment']; celda.border = estilos['border']
    
    ws.row_dimensions[1].height = 30

    # Analizar formatos de columnas basados en el nombre
    formatos_columna = {}
    for col_idx, nombre_col in enumerate(columnas, start=1):
        if 'EUR' in nombre_col:
            formatos_columna[col_idx] = '#,##0.00 €'
        elif '%' in nombre_col:
            formatos_columna[col_idx] = '0.00"%"'
        elif 'kg' in nombre_col or 'min' in nombre_col:
            formatos_columna[col_idx] = '#,##0.00'
        else:
            formatos_columna[col_idx] = 'General'

    # Datos
    for fila_idx, row in enumerate(df.values, start=2):
        fila_par = (fila_idx % 2 == 0)
        es_seccion = str(row[0]).startswith('--')
        
        for col_idx, valor in enumerate(row, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            if es_seccion:
                celda.font = Font(name=FUENTE_BASE, bold=True, color="FFFFFFFF")
                celda.fill = PatternFill("solid", fgColor=COLOR_SECCION)
            else:
                estilos = _estilo_dato(fila_par)
                celda.font = estilos['font']; celda.fill = estilos['fill']
                celda.border = estilos['border']
                celda.alignment = estilos['alignment']
                
                # Aplicar formato de número si es numérico
                if isinstance(valor, (int, float)):
                    celda.number_format = formatos_columna[col_idx]
                    
        ws.row_dimensions[fila_idx].height = 18

# =============================================================================
# ORQUESTADOR PRINCIPAL
# =============================================================================

def exportar_auditoria_excel(
    df_vuelos_crudo: pd.DataFrame,
    escenarios_dict: dict,
    df_dashboard: pd.DataFrame,
    df_slots: pd.DataFrame,
    params: dict,
    h_noreg: int,
    timeline: pd.DataFrame,
    r_min_newell: float,
    path: str,
    df_res_comprimido: pd.DataFrame = None,
    df_intermodal: pd.DataFrame = None, 
) -> None:
    """
    Construye el Excel consolidado para el reporte final.
    Organiza los escenarios GHP, Intermodales y Base de forma comparativa.
    """
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]

    # 1. EL DASHBOARD (Pestaña Reina para el Reporte)
    # -------------------------------------------------------------------------
    ws_dash = wb.create_sheet("00_Dashboard_KPIs", 0)
    _escribir_dataframe_con_formato(ws_dash, df_dashboard, COLOR_CABECERA_AZUL)
    _autoajustar_columnas(ws_dash)
    ws_dash.freeze_panes = "B2" # Congelar primera columna (Nombres de Escenario)

    # 2. ESCENARIOS DETALLADOS (Auditoría Vuelo a Vuelo)
    # -------------------------------------------------------------------------
    for nombre, df in escenarios_dict.items():
        if df is None or df.empty: continue
        
        color_tab = "1F4E79" # Azul
        if "GHP" in nombre: color_tab = "375623" # Verde
        if "Intermodal" in nombre: color_tab = "7030A0" # Púrpura
        
        ws_name = f"Audit_{nombre}"[:31] # Límite Excel 31 chars
        ws = wb.create_sheet(ws_name)
        ws.sheet_properties.tabColor = color_tab
        
        # Selección de columnas críticas para la auditoría (añadido co2_kg_vuelo si existe)
        cols_audit = [
            'ARCID','airline', 'Opr', 'ADEP','ATYP','distancia_km','minutes_eta',
            'assigned_slot','total_delay','air_delay','ground_delay','co2_kg_vuelo','flight_status'
        ]
        cols_finales = [c for c in cols_audit if c in df.columns]
        df_audit = df[cols_finales].sort_values('minutes_eta').reset_index(drop=True)
        
        _escribir_dataframe_con_formato(ws, df_audit, color_header="FF" + color_tab)
        _autoajustar_columnas(ws)
        ws.freeze_panes = "A2"

    # 3. COMPARATIVA INTERMODAL (WP4)
    # -------------------------------------------------------------------------
    if df_intermodal is not None and not df_intermodal.empty:
        ws_int = wb.create_sheet("WP4_Comparativa_Intermodal")
        ws_int.sheet_properties.tabColor = "7030A0"
        _escribir_dataframe_con_formato(ws_int, df_intermodal, COLOR_CABECERA_INT)
        _autoajustar_columnas(ws_int)

    # 4. PESTAÑAS DE INFRAESTRUCTURA Y PARÁMETROS (Soporte Técnico)
    # -------------------------------------------------------------------------
    _sheet_parametros(wb, params, h_noreg, r_min_newell)
    _sheet_matriz_slots(wb, df_slots)
    
    # 5. GUARDADO FINAL
    # -------------------------------------------------------------------------
    wb.save(path)
    print(f"✅ Excel de reporte generado y formateado: {os.path.basename(path)}")

# =============================================================================
# FUNCIONES DE APOYO
# =============================================================================

def _sheet_parametros(wb, params, h_noreg, r_min_newell):
    ws = wb.create_sheet("Config_Escenario")
    
    # Convertimos los params a un df sencillo para usar nuestra función de formato
    datos = [["Parámetro", "Valor"]]
    for k, v in params.items():
        datos.append([k, str(v)])
    datos.append(["h_noreg", str(h_noreg)])
    datos.append(["r_min_newell", str(r_min_newell)])
    
    df_params = pd.DataFrame(datos[1:], columns=datos[0])
    _escribir_dataframe_con_formato(ws, df_params, COLOR_CABECERA_AZUL)
    _autoajustar_columnas(ws)

def _sheet_matriz_slots(wb, df_slots):
    ws = wb.create_sheet("Config_Slots")
    _escribir_dataframe_con_formato(ws, df_slots.reset_index(drop=True), COLOR_CABECERA_AZUL)
    _autoajustar_columnas(ws)